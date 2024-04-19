from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

def add_value(beeline, megafon, tele2, filename):
    '''
    Добавить значения последней строкой
    '''
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.active
    new_row = ws.max_row + 1

    ws.append((
        datetime.now().strftime('%d.%m.%Y'),
        f'{datetime.now().hour}:00',
        megafon,
        tele2,
        beeline
    ))
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for c in range(1, 6):
        ws.cell(new_row, c).border = thin_border
        ws.cell(new_row, c).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(new_row, c).font = Font(name='Times New Roman', size=14)
    wb.save()